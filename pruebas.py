import fitz
import re
import openpyxl
from shipment import Shipment2
from itertools import chain


class EmptyPDFException(Exception):
    pass


cp_regex = re.compile(r'\nCP: [0-9]{4,}\n')

with fitz.open("prueba.pdf") as doc:
    page = doc[0].getText()
    # print(page_1)
    found_cps_lines = cp_regex.findall(page)
    packages = []
    # for cp in found_cps_lines:
    #     cp_index = str(page).rindex(cp)
    #     after_index = cp_index + len(cp)
    #     package = page[0:after_index]
    #     page = page[after_index:]
    #     packages.append(package)
    # print(packages)

    for cp in found_cps_lines:
        page = str(page).replace(cp, f'{cp}<<<COLUMN-END>>>\n')

    parts = page.split('<<<COLUMN-END>>>\n')

    for part in parts:
        print(part)

    lines = part[0].split("\n")

    shipment = Shipment2()
    if lines[0] == "":
        return None

    if lines[0].isnumeric():
        shipment.traking_id = self.without(lines[4], "Tracking: ")
        shipment.domicilio = self.without(lines[11], "Direccion: ")
        if "Referencia" in lines[-2]:
            shipment.entrecalles = self.without(lines[-2], "Referencia: ")
        shipment.codigo_postal = self.without(lines[-1], "CP: ")
        shipment.localidad = lines[8]
        shipment.partido = lines[7]
        shipment.destinatario = self.without(lines[9], "Destinatario: ")
        shipment.detalle_envio = "0-1"
    else:
        shipment.traking_id = self.without(lines[1], "Tracking: ")
        shipment.domicilio = self.without(lines[10], "Direccion: ")
        if "Referencia" in lines[-2]:
            shipment.entrecalles = self.without(lines[-2], "Referencia: ")
        shipment.codigo_postal = self.without(lines[-1], "CP: ")
        shipment.localidad = lines[7]
        shipment.partido = lines[6]
        shipment.destinatario = self.without(lines[8], "Destinatario: ")
        shipment.detalle_envio = "0-1"

    def without(self, s: str, replaceable: str):
        return s.replace(replaceable, "")


class CantConvertNothing(Exception):
    pass


class FileWithNoExtension(Exception):
    pass


class InvalidExtension(Exception):
    pass


class MLExtractor(object):

    def __init__(self, filename: str = None, **kwargs):
        self.filename = filename
        super().__init__(**kwargs)

    def convert(self, filepath: str = None) -> str:
        """Initializes convertion of file to csv

        Args:
            filepath (str, optional): file name. Defaults to None.

        Returns:
            csv: containing shipments's data
        """

        # If filepath was provided here, update self.file
        if filepath:
            # If filepath's not a str
            if type(filepath) != str:
                raise TypeError("Filepath must be a str.")
            self.file = filepath

        # If no filepath was provided
        if not self.file:
            raise CantConvertNothing("No filepath was provided.")

        try:
            # Find last '.' (stop)
            last_stop = self.file.rindex(".")
            # Get the extension
            extension = self.file[last_stop+1:]
        except ValueError:
            raise FileWithNoExtension("File didn't contain an extension.")

        if extension == 'pdf':
            # Do pdf work
            self.convert_pdf()
            pass
        elif extension == 'xlsx':
            # Do xlsx work
            print("do xlsx")
            pass
        elif extension == 'csv':
            # Do csv work
            print("do csv")
            pass
        else:
            raise InvalidExtension("Unknown file extension.")

        return 1


class ShipmentExractorModule():

    def __init__(self, filepath, **kwargs):
        self.filepath = filepath

    def extract_shipments(self):
        pass


class MercadoLibreSubmodule(ShipmentExractorModule):

    def __init__(self, pdf, **kwargs):
        self.pdf = pdf
        self.shipments = []

    def extract_shipments(self, pdf):
        # Filter last pages containing summary table
        pages = filter(
            lambda page: 'Despacha tu producto' not in page.getText(), pdf)
        # Map each page to a list o shipments
        list_of_shipments_lists = map(self.page_to_shipments, pages)
        # Flatten the list containing list of shipments
        return list(chain(*list_of_shipments_lists))

    def page_to_shipments(self, page):
        page_text = page.getText()
        cp_regex = re.compile(r'(\nCP: [0-9]{4,}\n)')
        page_text = cp_regex.sub(r'\1<<<COLUMN-END>>>\n', page_text)
        parts = page_text.split('<<<COLUMN-END>>>\n')
        return map(self.part_to_shipment, parts[:-2])

    def part_to_shipment(self, part) -> Shipment2:
        lines = part.split("\n")
        if lines[0] == "":
            return None

        # MercadoLibre's PDF labels got 2 different patterns
        is_type_1 = lines[0].isnumeric()

        # Define the Shipment2's properties
        traking_id = self.without(lines[4 if is_type_1 else 1], "Tracking: ")
        domicilio = self.without(lines[11 if is_type_1 else 10], "Direccion: ")
        codigo_postal = self.without(lines[-1], "CP: ")
        localidad = lines[8 if is_type_1 else 7]
        partido = lines[7 if is_type_1 else 6]
        destinatario = self.without(
            lines[9 if is_type_1 else 8], "Destinatario: ")
        detalle_envio = "0-1"
        entrecalles = self.without(
            lines[-2], "Referencia: ") if "Referencia" in lines[-2] else ""
        phone = ""

        # Return the shipment with the declared properties
        return Shipment2(traking_id, domicilio, codigo_postal,
                         localidad, partido, destinatario,
                         detalle_envio, entrecalles, phone)

    def without(self, s: str, replaceable: str):
        return s.replace(replaceable, "")


class TiendaNubeSubmodule():
    def __init__(self, filepath, **kwargs):
        super().__init__(filepath, **kwargs)

    def extract_shipments2(self, pdf):
        # Filter last pages containing summary table
        pages = filter(
            lambda page: 'Despacha tu producto' not in page.getText(), pdf)
        # Map each page to a list o shipments
        list_of_shipments_lists = map(self.page_to_shipments, pages)
        # Flatten the list containing list of shipments
        return list(chain(*list_of_shipments_lists))

    def extract_shipments(self, pdf):
        # Map each page to a list o shipments
        list_of_shipments_lists = map(self.page_to_shipments, pdf)
        return list(chain(*list_of_shipments_lists))

    def page_to_shipments(self, page):
        page_text = page.getText()
        text = page_text.replace("\n", "<<<")
        try:
            found = re.findall(r'Enviar a:<<<(.+?)Producto', text)
            return map(self.part_to_shipment, found)
        except AttributeError:
            raise EmptyPDFException("The PDF provided has no shipments in it.")

    def part_to_shipment(self, part):
        # Get destination's info
        destination = re.findall(r'<<<(.*),<<<Argentina', part)[0]
        destination_parts = destination.split("<<<")

        # Define the Shipment2's properties
        traking_id = ""
        domicilio = ""
        codigo_postal = destination_parts[-1].split(",")[-1].strip()
        localidad = ""
        partido = destination_parts[-1].split(",")[-3].strip()
        destinatario = re.search(r'(.*?)<<<', part).group(1)
        detalle_envio = ""
        entrecalles = re.findall(r'Notas del cliente:<<<(.*)<<<', part)
        phone = re.findall(r'Teléfono: \+(\d*)', part)

        try:
            localidad = destination_parts[-1].split(",")[-4].strip()
            domicilio = " ".join(destination_parts[:-1]).strip()
        except IndexError:
            if len(destination_parts) == 2:
                localidad = partido
                domicilio = destination_parts[0].strip()
            else:
                localidad = destination_parts[-2].strip()
                domicilio = "".join(
                    destination_parts[:-2]).replace("<<<", " ").strip()

        # Return the shipment with the declared properties
        return Shipment2(traking_id, domicilio, codigo_postal,
                         localidad, partido, destinatario,
                         detalle_envio, entrecalles, phone)


class PDFModule(ShipmentExractorModule):

    def __init__(self, filepath, **kwargs):
        super().__init__(filepath, **kwargs)

    def extract_shipments(self):
        with fitz.open(self.filepath) as pdf:
            first_page = pdf[0]
            fp_text = first_page.getText()
            is_mercado_libre = self.is_mercado_libre(fp_text)
            submodule = MercadoLibreSubmodule() if is_mercado_libre \
                else TiendaNubeSubmodule()
            return submodule.extract_shipments()

    def is_mercado_libre(self, text):
        return 'Mercado Envíos' in text \
            or 'Flex' in text or 'Recorta' in text


class ExcelModule(ShipmentExractorModule):
    def __init__(self, filepath, **kwargs):
        super().__init__(filepath, **kwargs)

    def extract_shipments(self):
        wb = openpyxl.load_workbook(self.filepath)
        wb.get_sheet_names()


class CSVModule(ShipmentExractorModule):
    def __init__(self, filepath, **kwargs):
        super().__init__(filepath, **kwargs)


class Extractor(object):

    def __init__(self, filepath: str = None, **kwargs):
        if filepath and type(filepath) != str:
            raise TypeError("Filepath must be a str.")
        self.filepath = filepath

    def get_shipments_as_csv_str(self, filepath: str = None) -> str:
        """Initializes convertion of file to csv

        Args:
            filepath (str, optional): file name. Defaults to None.

        Returns:
            csv: containing shipments's data
        """

        # If filepath was provided here, update self.file
        if filepath:
            # If filepath's not a str
            if type(filepath) != str:
                raise TypeError("Filepath must be a str.")
            self.filepath = filepath

        # If no filepath was provided
        if not self.filepath:
            raise CantConvertNothing("No filepath was provided.")

        try:
            # Find last '.' (stop)
            last_stop = self.filepath.rindex(".")
            # Get the extension
            extension = self.filepath[last_stop+1:]
            shipments = self.do_extraction(extension)
            return self.shipments_to_csv(shipments)
        except ValueError:
            raise FileWithNoExtension("File didn't contain an extension.")

    def do_extraction(self, extension):
        if extension == 'pdf':
            # Do pdf work
            module = PDFModule()
            return module.extract_shipments()
        elif extension == 'xlsx':
            # Do xlsx work
            module = ExcelModule()
            return module.extract_shipments()
        elif extension == 'csv':
            # Do csv work
            module = CSVModule()
            return module.extract_shipments()
        else:
            raise InvalidExtension("Unknown file extension.")

    def shipments_to_csv(self, shipments):
        print(shipments)
        return str(shipments)

    def convert_pdf_to_shipments(self):
        with fitz.open(self.file) as pdf:
            first_page = pdf[0]
            if self.is_mercado_libre(first_page.getText()):
                self.convert_mercado_libre(pdf)
            else:
                self.convert_tienda_nube(pdf)
        return ""

    def is_mercado_libre(self, text):
        return 'Mercado Envíos' in text \
            or 'Flex' in text or 'Recorta' in text


# with fitz.open("tiendanube.pdf") as pdf:
#     text = pdf[0].getText()
#     text = text.replace("\n", "<<<")
#     # text = 'gfgfdAAA1234ZZZuijjk'
#     # regex = re.compile(r'')

#     try:
#         found = re.findall(r'Enviar a:<<<(.+?)Producto', text)
#     except AttributeError:
#         # AAA, ZZZ not found in the original string
#         found = ''  # apply your error handling

#     # print("text", text, "\n\n")
#     # found = [item.replace("<<<", "\n") for item in found]
#     for part in found:
#         print(part, "\n")
#         try:
#             name = re.search(r'(.*?)<<<', part).group(1)
#             print("name", name)
#             phone = re.findall(r'Teléfono: \+(\d*)', part)
#             print("phone", phone)
#             hole_direc = re.findall(r'<<<(.*),<<<Argentina', part)
#             print("hole_direc", hole_direc)
#             entrecalles = re.findall(r'Notas del cliente:<<<(.*)<<<', part)
#             print("entrecalles", entrecalles)

#             destination = re.findall(r'<<<(.*),<<<Argentina', part)[0]
#             destination_parts = destination.split("<<<")

#             cp = destination_parts[-1].split(",")[-1].strip()
#             provincia = destination_parts[-1].split(",")[-2].strip()
#             partido = destination_parts[-1].split(",")[-3].strip()

#             try:
#                 ciudad = destination_parts[-1].split(",")[-4].strip()
#                 domicilio = " ".join(destination_parts[:-1]).strip()
#             except IndexError:
#                 if len(destination_parts) == 2:
#                     ciudad = partido
#                     domicilio = destination_parts[0].strip()
#                 else:
#                     ciudad = destination_parts[-2].strip()
#                     domicilio = "".join(
#                         destination_parts[:-2]).replace("<<<", " ").strip()

#             print("domicilio:", domicilio)
#             print("ciudad:", ciudad)
#             print("CP:", cp)
#             print("partido:", partido)
#             print("provincia:", provincia)
#             print(("\n\n"))

#         except AttributeError:
#             print("error")

import re
import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

for name in wb.get_sheet_names():
    sheet = wb.get_sheet_by_name(name)
    print(sheet)
    min_row = sheet.min_row
    min_col = sheet.min_column
    max_row = sheet.max_row
    max_col = sheet.max_column
    first_row_is_header = True

    # Check if first row does not have numbers in its cells
    # It will indicate first row are headers
    first_row_values = list(map(lambda x: x.value, sheet[min_row]))
    temp = "".join(first_row_values)
    first_row_is_header = not bool(re.search(r'\d', temp))
    if first_row_is_header:
        min_row += 1
    # print(min_row)

# def get_col_order(first_row_values):

    default_order = {
        'TRACKING ID MERCADO LIBRE': 0,
        'DOMICILIO': 1,
        'ENTRECALLES': 2,
        'CODIGO POSTAL': 3,
        'LOCALIDAD': 4,
        'PARTIDO': 5,
        'DESTINATARIO': 6,
        'DNI DESTINATARIO': 7,
        'TELEFONO DESTINATARIO': 8,
        'DETALLE DEL ENVIO': 9,
    }

    # order = default_order
    print(first_row_values)
    for i, value in enumerate(first_row_values):
        print(i+1, value)
        col = i + 1
        default_order[value] = col
        print(default_order[value])

    print(default_order)
        

    # default_order_old = {
    #     ["TRACKING ID MERCADO LIBRE", "TRACKING MERCADO LIBRE", "MERCADO LIBRE"]: 0,
    #     ["DOMICILIO","DOM","DIRECCION"]: 1,
    #     ["ENTRECALLES","OBSERVACIONES","NOTA", "NOTAS"]: 2,
    #     ["CODIGO POSTAL","POSTAL","CODIGO","COD POSTAL", "CODIGO POST","CODPOST","COD POST"]: 3,
    #     ["LOCALIDAD","CIUDAD","LOC", "CIU"]: 4,
    #     ["PARTIDO","PART","MUNICIPIO"]: 5,
    #     ["DESTINATARIO","DEST"]: 6,
    #     ["DNI DESTINATARIO", "DNI"]: 7,
    #     ["TELEFONO", "TELEFONO DESTINATARIO", "TEL"]: 8,
    #     ["DETALLE DEL ENVIO", "DETALLE"]: 9

    # }



